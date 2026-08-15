"""Populate data/ with REAL values, replacing the SEED_APPROXIMATE files.

Companion to fetch_data.py (which documents sources); this executes the loads
that can be automated. Inputs land in a local cache dir first so the parse is
re-runnable offline. Each writer stamps data_quality=PUBLISHED and the vintage
so the "no seed reaches production" test in build_zip2co2_web.py can enforce it.

Sources actually used (2023 reporting year):
  zip_subregion.csv     USEPA/power-profiler GitHub app/data/zip.csv — the exact
                        crosswalk behind EPA's own Power Profiler (eGRID2023,
                        June 2025). One row per (zip, utility) with subregion and
                        a Predominant-Utility flag, which beats the published
                        xlsx: it keeps the utility-level disambiguation that
                        decision (1) in tier2_factors.py wants.
  zip_utility.csv       Same file, preserved at (zip, utility) grain for the
                        supplier-candidate layer (wires utility + eiaid).
  egrid_subregion_factors.csv  eGRID2023 Summary Tables rev2 (Tables 1 + 2).
                        CO2e is EPA's own column, never recomputed (the CH4/N2O
                        lb-vs-GWh trap). Mix fractions -> percent. Grid gross
                        loss fraction -> percent. GWP set is FIT from the data:
                        argmin over {AR4 (25,298), AR5 (28,265)} of
                        |CO2 + a*CH4 + b*N2O - CO2e| summed across subregions.
Run:  .venv/bin/python zip2co2/build_data.py
"""

import csv
import io
import os
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
DATA = HERE / "data"
CACHE = HERE / "cache"          # gitignored; raw downloads for reproducibility
UA = {"User-Agent": "Mozilla/5.0 (personal carbon calculator; data build script)"}

ZIP_CSV_URL = "https://raw.githubusercontent.com/USEPA/power-profiler/master/app/data/zip.csv"
EGRID_SUMMARY_URL = "https://www.epa.gov/system/files/documents/2025-06/summary_tables_rev2.xlsx"
EGRID_YEAR = 2023


def fetch(url, name):
    CACHE.mkdir(exist_ok=True)
    p = CACHE / name
    if not p.exists():
        req = urllib.request.Request(url, headers=UA)
        with urllib.request.urlopen(req, timeout=120) as r:
            p.write_bytes(r.read())
        print(f"downloaded {name}: {p.stat().st_size:,} bytes")
    return p


# ---------------------------------------------------------------- zip crosswalk
def build_zip_tables():
    raw = fetch(ZIP_CSV_URL, "power_profiler_zip.csv").read_text(encoding="utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(raw)))
    by_zip = defaultdict(list)
    for r in rows:
        z = r["zip"].strip().zfill(5)
        by_zip[z].append(r)

    with open(DATA / "zip_subregion.csv", "w", newline="") as f:
        f.write("# Source: EPA Power Profiler crosswalk (USEPA/power-profiler GitHub, eGRID2023).\n")
        f.write("# Distinct subregions per ZIP, predominant utility's subregion first.\n")
        w = csv.writer(f)
        w.writerow(["zip", "subregion_1", "subregion_2", "subregion_3"])
        multi = 0
        for z in sorted(by_zip):
            rs = sorted(by_zip[z], key=lambda r: r.get("Predominant Utility") != "1")
            subs = []
            for r in rs:
                s = r["SUBRGN"].strip()
                if s and s not in subs:
                    subs.append(s)
            if len(subs) > 1:
                multi += 1
            w.writerow([z] + subs[:3] + [""] * (3 - min(3, len(subs))))
    print(f"zip_subregion.csv: {len(by_zip):,} ZIPs ({multi:,} span >1 subregion)")

    with open(DATA / "zip_utility.csv", "w", newline="") as f:
        f.write("# Source: EPA Power Profiler crosswalk (USEPA/power-profiler GitHub, eGRID2023).\n")
        f.write("# (zip, wires utility) grain — feeds zip_supplier_candidates for non-CCA rows.\n")
        w = csv.writer(f)
        w.writerow(["zip", "state", "eiaid", "utility_name", "subregion", "predominant"])
        for z in sorted(by_zip):
            for r in by_zip[z]:
                w.writerow([z, r["state"], r["eiaid"], r["UtilName"], r["SUBRGN"],
                            "TRUE" if r.get("Predominant Utility") == "1" else "FALSE"])
    print(f"zip_utility.csv: {sum(len(v) for v in by_zip.values()):,} rows")


# ---------------------------------------------------------------- eGRID factors
GWP_SETS = {"AR4": (25.0, 298.0), "AR5": (28.0, 265.0)}

def build_egrid():
    p = fetch(EGRID_SUMMARY_URL, "egrid_summary_tables.xlsx")
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)

    t1 = list(wb["Table 1"].iter_rows(values_only=True))
    t2 = list(wb["Table 2"].iter_rows(values_only=True))

    # Table 1 data rows: col1=acronym col2=name col3=CO2 col4=CH4 col5=N2O col6=CO2e ... col17=loss
    rates = {}
    for row in t1[4:]:
        if not row[1] or not isinstance(row[3], (int, float)):
            continue
        rates[str(row[1]).strip()] = {
            "name": str(row[2]).strip(), "co2": float(row[3]), "ch4": float(row[4]),
            "n2o": float(row[5]), "co2e": float(row[6]),
            # PRMS has no published loss; fall back to the U.S. figure below.
            "loss": float(row[17]) * 100.0 if isinstance(row[17], (int, float)) else None,
        }
    us_loss = rates.get("U.S.", {}).get("loss") or 4.2
    for r in rates.values():
        if r["loss"] is None:
            r["loss"] = us_loss

    # Which GWP set does EPA's CO2e column embed? Fit both, keep the better one.
    fit = {}
    for name, (g_ch4, g_n2o) in GWP_SETS.items():
        fit[name] = sum(abs(r["co2"] + g_ch4 * r["ch4"] + g_n2o * r["n2o"] - r["co2e"])
                        for r in rates.values())
    gwp_set = min(fit, key=fit.get)
    print(f"GWP fit residuals: { {k: round(v, 3) for k, v in fit.items()} } -> {gwp_set}")

    # Table 2 data rows: col1=acronym ... col5..15 = coal,oil,gas,other fossil,nuclear,
    # hydro,biomass,wind,solar,geothermal,other-unknown (fractions)
    mixes = {}
    for row in t2[3:]:
        if not row[1] or not isinstance(row[5], (int, float)):
            continue
        (coal, oil, gas, ofossil, nuclear, hydro,
         biomass, wind, solar, geo, ounknown) = [float(x or 0) * 100 for x in row[5:16]]
        mixes[str(row[1]).strip()] = {
            "pct_coal": coal, "pct_oil": oil, "pct_natural_gas": gas,
            "pct_nuclear": nuclear, "pct_hydro": hydro, "pct_biomass": biomass,
            "pct_wind": wind, "pct_solar": solar, "pct_geothermal": geo,
            "pct_other": ofossil + ounknown,
        }

    mix_cols = ["pct_coal", "pct_natural_gas", "pct_oil", "pct_nuclear", "pct_hydro",
                "pct_wind", "pct_solar", "pct_geothermal", "pct_biomass", "pct_other"]
    with open(DATA / "egrid_subregion_factors.csv", "w", newline="") as f:
        f.write("# Source: EPA eGRID2023 Summary Tables rev2 (June 2025), Tables 1 and 2.\n")
        f.write("# CO2e is EPA's published ANNUAL TOTAL OUTPUT column (not non-baseload, never\n")
        f.write("# recomputed from CH4/N2O). Grid gross loss and mix converted fraction->percent.\n")
        f.write(f"# gwp_set determined by fit against EPA's own CO2e column (see build_data.py).\n")
        w = csv.writer(f)
        w.writerow(["egrid_year", "subregion", "subregion_name", "co2e_lb_per_mwh",
                    "co2_lb_per_mwh", "grid_gross_loss_pct"] + mix_cols
                   + ["gwp_set", "data_quality"])
        subs = sorted(set(rates) & set(mixes) | {s for s in rates if s in ("U.S.",)})
        n = 0
        for s in sorted(set(list(rates) + list(mixes))):
            r, m = rates.get(s), mixes.get(s)
            if not r or not m:
                print(f"  skipping {s}: only in {'Table 1' if r else 'Table 2'}")
                continue
            key = "US" if s in ("U.S.", "US") else s
            name = r["name"] if r["name"] not in ("None", "") else "U.S. average"
            w.writerow([EGRID_YEAR, key, name, round(r["co2e"], 3), round(r["co2"], 3),
                        round(r["loss"], 2)] + [round(m[c], 2) for c in mix_cols]
                       + [gwp_set, "PUBLISHED"])
            n += 1
    print(f"egrid_subregion_factors.csv: {n} subregions")


# ------------------------------------------------------------- residual mix
# Carbon-free flags per eGRID mix category under the "conventional" definition
# (matches fuel_factors.csv: nuclear, hydro, wind, solar, geothermal, biomass).
CONVENTIONAL_CF_COLS = ("pct_nuclear", "pct_hydro", "pct_wind", "pct_solar",
                        "pct_geothermal", "pct_biomass")

def build_residual():
    """CRS 2025 Residual Mix (2023 data), captured to cache/ via browser — the
    site's Cloudflare blocks scripted fetch; see the JSON's provenance note.

    Two derivations on top of the published table:
    - includes_ch4_n2o: the published system rate equals eGRID2023 CO2e exactly
      (AKGD 905.109 etc.), so this release is CO2e-based -> TRUE, no rescale.
    - pct_carbon_free_residual is NOT published. Derived with the same
      arithmetic CRS uses for the rate: voluntary REC MWh are carbon-free by
      construction, so residual CF share = (CF_gen - voluntary) / (netgen -
      voluntary), with CF_gen = conventional-definition share x net generation
      from eGRID Table 2. Floored at 0 (a subregion can sell more RECs than
      one category's generation, not more than its total CF generation).
    """
    import json
    crs = json.loads((CACHE / "crs_residual_2025.json").read_text())

    wb = openpyxl.load_workbook(CACHE / "egrid_summary_tables.xlsx",
                                read_only=True, data_only=True)
    netgen, cf_share = {}, {}
    for row in wb["Table 2"].iter_rows(min_row=4, values_only=True):
        if not row[1] or not isinstance(row[4], (int, float)):
            continue
        sub = str(row[1]).strip()
        (coal, oil, gas, ofossil, nuclear, hydro,
         biomass, wind, solar, geo, ounknown) = [float(x or 0) for x in row[5:16]]
        netgen[sub] = float(row[4])
        cf_share[sub] = nuclear + hydro + biomass + wind + solar + geo

    with open(DATA / "greene_residual_mix.csv", "w", newline="") as f:
        f.write("# Source: CRS 2025 Residual Mix Emissions Rates (2023 data),\n")
        f.write("# resource-solutions.org/2025-residual-mix/ ('Adjusted System Mix', 12-month vintage).\n")
        f.write("# System rates match eGRID2023 CO2e exactly -> series is CO2e; includes_ch4_n2o=TRUE.\n")
        f.write("# pct_carbon_free_residual derived: (CF_gen - voluntary RE MWh)/(netgen - voluntary),\n")
        f.write("# conventional carbon-free definition, eGRID Table 2 generation. See build_data.py.\n")
        w = csv.writer(f)
        w.writerow(["residual_year", "subregion", "residual_lb_per_mwh", "includes_ch4_n2o",
                    "pct_carbon_free_residual", "source_egrid_vintage", "data_quality"])
        for r in crs["rows"]:
            sub = r["sub"]
            ng, vol = netgen.get(sub), r["vol_mwh"]
            if ng and ng > vol:
                cf_res = max(0.0, (cf_share[sub] * ng - vol) / (ng - vol)) * 100.0
            else:
                cf_res = cf_share.get(sub, 0.0) * 100.0
            w.writerow([crs["residual_year"], sub, round(r["resid"], 3), "TRUE",
                        round(cf_res, 2), EGRID_YEAR, "PUBLISHED"])
        # US row: generation-weighted mean of subregion residuals for the ZIP-miss path.
        tot = sum(netgen.get(r["sub"], 0) for r in crs["rows"])
        us_rate = sum(r["resid"] * netgen.get(r["sub"], 0) for r in crs["rows"]) / tot
        us_cfg = sum(cf_share.get(r["sub"], 0) * netgen.get(r["sub"], 0) for r in crs["rows"])
        us_vol = sum(r["vol_mwh"] for r in crs["rows"])
        us_cf = max(0.0, (us_cfg - us_vol) / (tot - us_vol)) * 100.0
        w.writerow([crs["residual_year"], "US", round(us_rate, 3), "TRUE",
                    round(us_cf, 2), EGRID_YEAR, "DERIVED_WEIGHTED_MEAN"])
    print(f"greene_residual_mix.csv: {len(crs['rows'])} subregions + US")


if __name__ == "__main__":
    build_zip_tables()
    build_egrid()
    build_residual()
