"""Build gridcarbon/data/zip_ba.csv — national ZIP -> balancing authority.

Join: ZIP -> wires utility eiaid (EPA Power Profiler crosswalk, predominant
utility first) -> BA code (EIA-861 Sales_Ult_Cust, utility x state, picking the
BA with the most residential customers where a utility spans several).

ZIPs whose utility has no EIA-861 BA row (munis inside another BA's footprint,
PR/AK/HI islands without EIA-930 hourly data) simply drop out — the app falls
back to the flat eGRID subregion factor from the v1 (zip2co2) dataset.

Run:  ../.venv/bin/python make_zip_ba.py
"""

import csv
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
ZIP_UTIL = os.path.join(ROOT, "zip2co2", "data", "zip_utility.csv")
EIA861 = os.path.join(ROOT, "zip2co2", "cache", "eia861", "Sales_Ult_Cust_2023.xlsx")
OUT = os.path.join(HERE, "gridcarbon", "data", "zip_ba.csv")


def utility_ba_map():
    """(eiaid, state) -> BA code, by max residential customers; eiaid -> modal BA."""
    wb = openpyxl.load_workbook(EIA861, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=4, values_only=True)
    best = {}
    by_util = defaultdict(lambda: defaultdict(float))
    for r in rows:
        try:
            uid, state, ba = str(int(r[1])), str(r[6]).strip(), str(r[8] or "").strip()
        except (TypeError, ValueError):
            continue
        if not ba or ba == "None":
            continue
        try:
            cust = float(r[11])   # "." = EIA's withheld marker
        except (TypeError, ValueError):
            cust = 0.0
        key = (uid, state)
        if cust >= best.get(key, (None, -1))[1]:
            best[key] = (ba, cust)
        by_util[uid][ba] += cust
    modal = {uid: max(bas.items(), key=lambda kv: kv[1])[0] for uid, bas in by_util.items()}
    return {k: v[0] for k, v in best.items()}, modal


def main():
    by_state, modal = utility_ba_map()
    n, miss = 0, 0
    seen = set()
    with open(OUT, "w", newline="") as f:
        f.write("# ZIP -> BA: EPA Power Profiler wires utility (predominant first) joined to\n")
        f.write("# EIA-861 Sales_Ult_Cust BA codes (max residential customers per utility-state).\n")
        f.write("# Missing ZIPs fall back to the flat eGRID subregion factor (v1 dataset).\n")
        w = csv.writer(f)
        w.writerow(["zip", "state", "ba", "utility", "load_shape_id"])
        rows = list(csv.DictReader(ln for ln in open(ZIP_UTIL)
                                   if not ln.startswith("#")))
        rows.sort(key=lambda r: (r["zip"], r["predominant"] != "TRUE"))
        for r in rows:
            z = r["zip"]
            if z in seen:
                continue
            ba = by_state.get((r["eiaid"], r["state"])) or modal.get(r["eiaid"])
            if ba is None:
                miss += 1
                continue
            seen.add(z)
            w.writerow([z, r["state"], ba, r["utility_name"], ba])
            n += 1
    print(f"zip_ba.csv: {n:,} ZIPs mapped ({miss:,} rows unmapped -> subregion fallback)")


if __name__ == "__main__":
    main()
